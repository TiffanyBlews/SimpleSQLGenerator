from openpyxl import load_workbook
import os.path
pwd = os.path.dirname(__file__)
book = load_workbook(pwd+'\数据库设计表.xlsx', data_only=True)

IS_MYSQL = False

res = []
for i, sheet in enumerate(book.sheetnames):
    sql = f'''create table {sheet}
(
'''
    for j, row in enumerate(book[sheet].values):
        line = ''

        if j > 0 and any(row):
            for k, cell in enumerate(row):

                if k in [0, 2] and cell:
                    line += cell + ' '
                elif k == 1 and cell:
                    if IS_MYSQL:
                        line += 'comment \'%s\' ' % cell
                elif k == 3 and cell == 'Y':
                    line += 'primary key '
                elif k == 4 and cell == 'Y' and not line.endswith('primary key '):
                    line += 'unique '
                elif k == 5 and cell == 'N':
                    line += 'not null '
                elif k == 6:
                    line += cell if cell != None else ''
                    line += ',\n'
                if cell == None:
                    continue
        sql += line
    sql += ');\n\n'
    res.append(sql)

with open(pwd+'\output.sql', 'w+', encoding='utf8') as f:
    f.write(''.join(s for s in res))

print("\n\n6\n\n")
