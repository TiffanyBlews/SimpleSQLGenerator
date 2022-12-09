from openpyxl import load_workbook
import os.path
pwd = os.path.dirname(__file__)
book = load_workbook(pwd+'\数据示例.xlsx', data_only=True)

IS_MYSQL = False
res=[]
for i,sheet in enumerate(book.sheetnames):
    sql=f'''insert into {sheet}
( '''
    cols= [c for c in next(book[sheet].values)]
    cols=list(filter(None,cols))
    sql+= ', '.join(cols)
    sql+=' )\nvalues\n'
    input=[]

    for j, row in enumerate(book[sheet].values):
        if j>0 and any(row):
            line='('
        
            for k, cell in enumerate(row):
                if k>= len(cols):
                    break
                if cell is not None:
                    line+='\''+str(cell)+'\', '
                else:
                    line+='NULL,'
            line+='),\n'
            sql+=line
    res.append(sql+'\n')

# print(res)
with open(pwd+'\outputdata.txt', 'w+', encoding='utf8') as f:
    f.write('\n'.join(s for s in res))

print("\n\n6\n\n")
